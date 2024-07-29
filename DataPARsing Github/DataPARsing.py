import pandas as pd
import numpy as np
import openpyxl
import fsspec
import time

import matplotlib.pyplot as plt
import matplotlib as mpl

from dateutil.parser import parse
from scipy.signal import savgol_filter, butter, lfilter, freqs

import pickle

from copy import deepcopy

import os
import os.path as osp

default_sampling_interval = 10 #mins

color_list = [[0.5, 0.0, 0],
[0.8333333333333333, 0.0, 0],
[0.5, 0.5, 0],
[0.8333333333333333, 0.5, 0],
[0.5, 0.0, 1],
[0.8333333333333333, 0.0, 1],
[0.5, 0.5, 1],
[0.8333333333333333, 0.5, 1]]
number_of_series = 0

def filename_from_directory():
    ls = [f for f in os.listdir() if osp.isfile(f) and "PICKLE" in f]
    for i, f in enumerate(ls):
        print(f"({i})\t{f}")
    return ls[int(input())]

class Stopwatch:
    def __init__(self, name=None):
        self.start_time = time.time()
        self.laps = 0
        self.name = name
    def stop(self,verbose = True):
        self.end_time = time.time()
        time_elapsed = self.end_time - self.start_time
        if verbose:
            if self.name: print("TIMER:",self.name,end="\t")
            print("LAP:",self.laps)
            print("Time Elapsed:\t",time_elapsed)

        self.laps +=1
        return time_elapsed
    
def outside(val, args):
    avg = args[0]
    std = args[1]
    z = args[2]
    return (val > avg + z*std or val < avg - z*std)
def under(val, args):
    return val<args
def over(val, args):
    return val>args
def not_within(val,args):
    #args must be a list or tuple
    return (val<args[0] or val>args[1])
def within(val,args):
    return (val>args[0] and val<args[1])

class PARC_TimeSeries:
    linsym = "-"
    def __init__(self, xlfilename, sheetname,
                 tag =None, tag_description=None, tag_units=None,
                 pull_data=True,color=None, verbose = False, version = 0):
        if verbose:sw = Stopwatch("TS TIMER")
        if pull_data:
            if version ==0: self.series = getTimeSeries(xlfilename,sheetname)
            if version ==1: self.series = getTimeSeries_v2(xlfilename,data_col=sheetname) #in this case, pass the column number instead of the sheetname when initialization the PARC object
        
        self.tag = tag
        self.tag_description = tag_description
        self.tag_units = tag_units
        
        td = self.tag_description
        self.acronym = td[0:3]
        for i,_ in enumerate(td): # make an acronym
            if i<3: continue
            if td[i-1] == " ": self.acronym += td[i]

        if color==None:
            global number_of_series
            global color_list
            number_of_series += 1
            self.color = color_list[number_of_series%len(color_list)]
        else: self.color = color
        if verbose:sw.stop()
    def get_average(self):
        mean = np.mean(self.series)
        if type(mean) == float: return mean
        return self.series.mean()
         
    def get_standardDeviation(self):
        return np.std(self.series)
    
    def plot(self, ax, auto_title=True, scale=1, y_offset = 0, lintype = None):
        if lintype == None: lintype = self.linsym
        ax.plot(
            (self.series + y_offset) * scale,
            lintype,
            label = self.tag_description,
            color = self.color)
        ax.legend()
        if auto_title:
            ax.set_title(self.tag +  "\n" + self.tag_description)
            ax.set_ylabel(self.tag_units)
            ax.set_frame_on(1)
        
    def smooth(self,windowsize=10,polynomialorder=3):
        newPARCts = PARC_TimeSeries(None, None,
            tag = self.tag + " (smoothed)",
            tag_description = self.tag_description,
            tag_units = self.tag_units,
            pull_data=False)
        newPARCts.series = seriesSVFilter(self.series, windowsize, polynomialorder)
        return newPARCts
    
    def __sub__(self,other): # should contract the setup into another little function but that's a job for later!
        opsym = " - "
        newPARCts = PARCts_arbitration(self,other,opsym)
        if type(other) == PARC_TimeSeries: newPARCts.series = np.subtract(self.series,other.series)
        else: newPARCts.series = self.series - other
        return newPARCts
    
    def __add__(self,other):
        opsym = " + "
        newPARCts = PARCts_arbitration(self,other,opsym)
        if type(other) == PARC_TimeSeries: newPARCts.series = np.add(self.series,other.series)
        else: newPARCts.series = self.series + other
        return newPARCts
    
    def __mul__(self,other):
        opsym = " x "
        newPARCts = PARCts_arbitration(self,other,opsym)
        if type(other) == PARC_TimeSeries: newPARCts.series = np.multiply(self.series,other.series)
        else: newPARCts.series = self.series * other
        return newPARCts
    
    def __truediv__(self,other):
        opsym = " / "
        newPARCts = PARCts_arbitration(self,other,opsym)
        if type(other) == PARC_TimeSeries: newPARCts.series = np.divide(self.series,other.series)
        else: newPARCts.series = self.series / other
        return newPARCts
    def __pow__(self,other):
        deg = str(other)
        newPARCts = PARC_TimeSeries(None, None,
            tag = self.tag + " to the " + deg + "th",
            tag_description = self.tag_description + " to the " + deg + "th",
            tag_units = self.tag_units + " to the " + deg + "th",
            pull_data=False)
        newPARCts.series = np.power(self.series, other)
        return newPARCts
    
    def slope(self):
        newPARCts = PARC_TimeSeries(None, None,
            tag = self.tag + " (per min)",
            tag_description = self.tag_description + " (slope)",
            tag_units = self.tag_units + " (per min)",
            pull_data=False)
        newPARCts.series = seriesDerivative(self.series)
        return newPARCts
    def round(self,width=1):
        newPARCts = PARC_TimeSeries(None, None,
            tag = self.tag + " (bucket size: " + str(width) + " )",
            tag_description = self.tag_description + " (bucket size: " + str(width) + " )",
            tag_units = self.tag_units + " (bucket size: " + str(width) + " )",
            pull_data=False)
        newPARCts.series = np.round( (100/width) * self.series,-2) *  (width/100)
        return newPARCts
    
    def without_outliers(self):
        avg = self.get_average()
        std = self.get_standardDeviation()
        lower_bound = avg - 3*std
        upper_bound = avg + 3*std
        newPARCts = PARC_TimeSeries(None, None,
            tag = self.tag  + " (cleaned)",
            tag_description = self.tag_description + " (cleaned)",
            tag_units = self.tag_units,
            pull_data=False)
        newPARCts.series = deepcopy(self.series)
        for i, datum in enumerate(newPARCts.series):
            if datum < lower_bound or datum > upper_bound: newPARCts.series.iloc[i] = np.nan
        return newPARCts
    
    def normalize(self): #center about y=0, then shrink/stretch to std = 1
        newPARCts = PARC_TimeSeries(None, None,
            tag = self.tag + " (normalized)",
            tag_description = self.tag_description + " (normalized)",
            tag_units = self.tag_units,
            pull_data=False)
        newPARCts.series = (self.series - self.get_average()) / self.get_standardDeviation()
        print( "AVG:",self.get_average(), "\nSTD:",self.get_standardDeviation())
        return newPARCts
    def cleanse_timestamps(self, events):
        tpl = events.ess #timestamp pair list
        newPts = deepcopy(self)
        i = 0
        for t in newPts.series.index:
            if t >= tpl[i][0]: newPts.series[t] = np.NaN
            if t > tpl[i][1]:
                i += 1
                if i == len(tpl): break
        return newPts
    
    def keep_timestamps(self, events):
        tpl = events.ess #timestamp pair list
        newPts = deepcopy(self)
        i = 0
        for t in newPts.series.index:
            if t <= tpl[i][0]: newPts.series[t] = np.NaN
            if t > tpl[i][1]:
                i += 1
                if i == len(tpl): break
        return newPts
    
    def get_event_timestamps(self, #get start and end times of 'events' where the value proves some function true
                          eval_funct = outside, #by default, find values outside 2 stdevs from mean... ie outliers
                          eval_funct_arg = "get avg and std 2"):
        if type(eval_funct_arg) == str:
            if eval_funct_arg[0:15] == "get avg and std":
                if eval_funct_arg[-1].isnumeric():
                    num = float(eval_funct_arg[15:])
                    print(num)
                else: num = 2
                eval_funct_arg = [self.get_average(), self.get_standardDeviation(), num]
                
        print(eval_funct_arg)
        events = []
        for i, val in enumerate(self.series): # get all indices for which eval_funct(val) is True
            if eval_funct( val, eval_funct_arg ): events.append( i )
        if len(events) < 2: return None

        ser = self.series
        ess = []
        start = ser.index[events[0]]
        for i, t in enumerate(events[1:]): # get start/stop index pairs from events[]
            #note: because enumerating events[1:], t = events[i+1]
            if t!= events[i] + 1: #if this value is 1 bigger than prev value, we are continuous. 
                ess.append( [start,ser.index[events[i]] ]) #if not, it is a new event and the prev value is the end of the prev event
                start = ser.index[t]
        if ess[-1][1] != t: #edge case: last event
            ess.append( [ start, ser.index[t] ] )
        ess = Events(ess)
        return ess
    


    def corr(self,other):
       return self.series.corr(other.series)

class Events: # a list of pairs of timestamps describing starts and ends of events. Works also as a Pandas Series.
    def __init__(self,ess):
        self.ess = ess
        self.update_series()
        
    @property
    def true_duration(self):
        perc = self.true_duration_percent
        return perc * self.total_duration
    @property
    def true_duration_percent(self):
        return np.mean(self.series)
    @property
    def total_duration(self):
        return self.series.index[-1] - self.series.index[0]
        
    def pad(self, mins): #exand or contract the size of each event by x mins on each end
        #alternatively, pass a tuple to pad by x minutes before and y minutes after
        if type(mins) == list or type(mins) == tuple:
            mins_before = str(mins[0]) + "m"
            mins_after = str(mins[1]) + "m"
        else:
            mins_before = mins_after = str(mins) + "m"
            
        mins_before = pd.Timedelta(mins_before)
        mins_after = pd.Timedelta(mins_after)
        
        self.ess = [ [pair[0] - mins_before, pair[1] + mins_after] for pair in self.ess]
        self.merge_overlaps()
        self.update_series()
    
    def nudge(self, mins): #shift to the future x mins
        mins = str(mins) + "m"
        mins = pd.Timedelta(mins)
      
        self.ess = [ [pair[0] + mins, pair[1] + mins] for pair in self.ess]
        self.merge_overlaps()
        self.update_series()
    def plot(self, ax, scale=1, y_offset = 0, lintype = None):
        if lintype == None: lintype = self.linsym
        for [s,e] in self.ess:
            ax.axvspan(s,e,color='red',alpha=0.2)
        
    def merge_overlaps(self):
        i = 0
        while i < len(self.ess) -1:
            if self.ess[i][1] > self.ess[i+1][0]:
                self.ess[i][1] = self.ess[i+1][1]
                del self.ess[i+1]
            else: i+=1
    def combine(self, other, logic = "AND"): #currently no room for other logic, lol
        if type(other) != type(self): raise(TypeError)
        if logic == "AND": thresh = 1
        elif logic == "OR": thresh = 0
        #oh this is going to be horrendously innefficient...
        #steal code from get_event_timestamps
        newseries = ns = np.add(self.series ,other.series)
        ess = []
        start = 0
        for t, val in zip(ns.index, ns.values):
            if start == 0: # event hasn't started
                if val>thresh: start = t #catch AND condition
            else: #start has been caught
                if val<=thresh:  #catch NAND condition
                    end = t 
                    ess.append([start,end]) #add timestamp pair
                    start = 0 # look for start again...
        if start != 0: # catch straggler if there is one
            end = t 
            ess.append([start,end]) #add timestamp pair
        ess = Events(ess)
        return ess
        
            
    def update_series(self):
        self.linsym = "-"
        
        start_t = self.ess[0][0]
        end_t = self.ess[-1][-1]
        timelist = [start_t]
        values = [True]
        while(timelist[-1] < end_t):
            timelist.append( timelist[-1] + pd.Timedelta('1m'))
            values.append(True)
            
        self.series = pd.Series(data=values, index = timelist )

        i = 0
        for t in self.series.index:
            if t <= self.ess[i][0]: self.series[t] = False
            if t > self.ess[i][1]:
                i += 1
                if i == len(self.ess): break

    def getDeltas(self, Pts, return_format = list, percents = False):
        s = deepcopy(Pts.series)
        if percents: deltas = [ s[tpair[1]] / s[tpair[0]] - 1 for tpair in self.ess] #end/start - 1
        else: deltas = [ s[tpair[1]] - s[tpair[0]] for tpair in self.ess]
        
        
        if return_format == list: return deltas
        elif return_format == Events:
            retEvents = deepcopy(self)
            
            i = 0
            for t in retEvents.series.index:
                if t <= self.ess[i][0]: retEvents.series[t] = False
                else: retEvents.series[t] = deltas[i]
                if t > self.ess[i][1]:
                    i += 1
                    if i == len(self.ess): break
            return retEvents
        elif return_format == pd.Series:
             return pd.Series(deltas)
 
class PARC_tsBuckets:
    def __init__(self,PARCts,width):
        PARCts = PARCts.round(width) #round
        self.series = ts = PARCts.series[~np.isnan(PARCts.series)] #remove nans
        self.bucket_vals = bv = sorted(set(ts.values)) #get values
       
        self.bucket_arr = [deepcopy(PARCts) for val in bv]
        for i, val in enumerate(bv):
            (self.bucket_arr[i]).series = ( ts[(ts==val)] )
            (self.bucket_arr[i]).linsym = "x"
        
    
def best_corr_weighted_sum(ts1, ts2, tscomp, precision = 3, verbose = True):
    t = Stopwatch()
    weight = 0
    step_size = ts1.get_average() / ts2.get_average() / 5
    direction = 1
    # figure out best direction to start
    if (ts1 + ts2*0.001).corr(tscomp) > (ts1 - ts2*0.001).corr(tscomp): direction = 1
    else: direction = -1
    # keep adding to weight until results get worse.
    #also shut down the function if it runs for too long
    for i in range(0,precision):
        prev_corr = (ts1 + ts2*weight).corr(tscomp)
        weight += step_size*direction
        while prev_corr < (ts1 + ts2*weight).corr(tscomp):
            prev_corr = (ts1 + ts2*weight).corr(tscomp)
            weight += step_size*direction
            if verbose: print(prev_corr,"\t\t", weight, step_size, direction)
            if t.stop(False) > precision / 2:
                print(weight, prev_corr)
                return 0
        direction *= -1 #other direction
        step_size *= 1/16 #smaller steps
    #once results get worse, reverse direction with a smaller step size
    print(weight, prev_corr)
    return weight

def bcws2(ts1, ts2, tscomp, precision = 3, verbose = True):
    width = w = 1000
    center = c = 0
    for i in range(precision):
        weights = r = np.linspace(c-w,c+w,50)
        corr_arr = cl = [ (ts1 + ts2 * w).corr(tscomp) for w in r]
        #print(weights)
        #update center
        newc = r[ cl.index( max(cl) ) ]
        if newc <= c-w or newc >= c+w: w*= 10
        else: w *= 0.1
        c = newc
        if verbose: print(c)
    return c
        

def PARCts_arbitration(sts, ots, opsym): #self, other, operation symbol (" + "," - ",...)
    if type(ots) == PARC_TimeSeries: tag_strings = [
            sts.tag +  opsym + ots.tag,
            sts.tag_description + opsym + ots.tag_description,
            colcom( sts.color, ots.color)
            ]
    else: tag_strings = [
            sts.tag +  opsym + str(ots),
            sts.tag_description + opsym + str(ots),
            colcom( sts.color, [0.2,0.2,0.2])
            ]
    return PARC_TimeSeries(None, None,
            tag = tag_strings[0],
            tag_description = tag_strings[1],
            tag_units = None,
            color = tag_strings[2],
            pull_data=False) 
def rawdataread(filename, sheetname,time_col = 0,data_col = 1, skiprows=16): #wrapper for pd.read_excel that's good for Raw Data from dataPARC
    return pd.read_excel(filename,
                         engine='openpyxl',
                   sheet_name=sheetname,
                   header =None,
                   skiprows = skiprows, # this is just the format of "Raw Data" in dataPARC excel...
                   usecols=[time_col, data_col], #sometimes they go into cols 2,3 but I haven't the faintest why they do this
                   names =['Timestamps','Values']
                   )
def getTimeSeries_v2(filename, data_col,sheetname="Data Normalize"): #get the raw data from XL into a
    rawdata = rawdataread(filename, sheetname,time_col=3,data_col= data_col, skiprows = 18)
    series = pd.Series(list(rawdata.Values), index=rawdata.Timestamps )
    #print(series1)
    return series

def getTimeSeries(filename, sheetname, time_interval_string = str(default_sampling_interval)+"min"): #get the raw data from XL into a
                                                                        #time series with a constant time step
                                                                        #which is good for comparisons
    rawdata = rawdataread(filename, sheetname)
    series = pd.Series(list(rawdata.Values), index=rawdata.Timestamps )
    #print(series)
    #resample to 10 mins
    series = series.resample(time_interval_string).mean()
    #print(series1)
    return series


def getData_v2(filename):
    xlfile = filename
    xlfile_wb = openpyxl.load_workbook(xlfile)
    ws = xlfile_wb["Data Normalize"]
    tags = [c.value for c in ws[1][4:]]
    descriptions = [c.value for c in ws[2][4:]]
    units = [c.value for c in ws[3][4:]]

    data_list = [ PARC_TimeSeries(xlfile, i+4, tags[i], descriptions[i],units[i],version = 1) for i, _ in enumerate(tags)]
    data = { str(data.tag):data for data in data_list }
    for tag, description, unit in zip(tags, descriptions, units):
        print("\t\t\t",tag,"\n\tDESC:\t",description,"\n\tUNITS:\t",unit,"\n")

    data_file = open('data_pickle','ab')
    pickle.dump( data, data_file)
    data_file.close()
    
    return data
    
def getData(filename):
    sw = Stopwatch("Get Metadata")
    xlfile = filename
    xlfile_wb = openpyxl.load_workbook(xlfile)

    sw.stop()
    
    #it is important that the PARC sheet (or any non-data sheet) is the LAST PAGE!!!!
    #it is also important that the last page is called 'PARC'
    sheetnames = [sheet.title for sheet in xlfile_wb]
    sheetnames.pop() #remove last sheet (PARC or whatever)

    units = [xlfile_wb[sheetname]['B3'].value for sheetname in sheetnames]
    sw.stop()
    descriptions = [xlfile_wb[sheetname]['B2'].value for sheetname in sheetnames]
    sw.stop()
    #this line only works on a LOT of assumptions about setup style
    # 1) info page is called 'PARC'
    # 2) Data Raw (x) corresponds to column x
    # this is acheivable if you create the new Data Raw sheets in order left to right
    #    on PARC and right to left on the bottom tab navigator in Excel
    tags = [xlfile_wb['PARC'].cell(row=2,column=(i+1)).value for i, _ in enumerate(sheetnames)]
    tags.reverse() #a convention which may break some things!
    sw.stop()
    
    data_list = [ PARC_TimeSeries(xlfile, sheetnames[i], tags[i], descriptions[i],units[i]) for i, _ in enumerate(sheetnames)]
    data = { str(data.tag):data for data in data_list }
    for tag, description, unit in zip(tags, descriptions, units):
        print("\t\t\t",tag,"\n\tDESC:\t",description,"\n\tUNITS:\t",unit,"\n")
    
    return data

def getTimeString():
    t = time.localtime()
    return str(t.tm_year) + "-" + str(t.tm_mon) + "-" + str(t.tm_mday) + "_" + str(t.tm_hour * 100 + t.tm_min)

def getDataQuick(filename, picklename=getTimeString() + "_PICKLE"):
    if input("HAVE YOU MODIFIED THE XL FILE \nSINCE THE LAST RUN, OR IS THIS \nYOUR FIRST RUN? (y/n):") == "y":
        if input("IS YOUR WORKBOOK MANY SHEETS OR ONE PAGE? (m/o)") == "m":
            data = getData(filename)
        else: data = getData_v2(filename)
        data_file = open(filename[:-5]+"_"+picklename,'ab')
        pickle.dump( data, data_file)
        data_file.close()
    else:
        picklefilename = filename_from_directory()
        #input("\nInput The Name of the Pickle File...\n")
        data_file = open(picklefilename,'rb')
        data = pickle.load(data_file)
        data_file.close()
    return data

def seriesDerivative(series): #take the derivative of a time series, return a time series
    return pd.Series(np.diff(series),index=series.index[1:])#np.diff returns 1 less values, so reduce the time series size by one

def seriesSVFilter(series, windowsize, polynomialorder):
    return pd.Series(savgol_filter(series, windowsize, polynomialorder), index=series.index)



def butter_lowpass(cutoff, fs, order=5):
    return butter(order, cutoff, fs=fs, btype='low', analog=False)

def butter_lowpass_filter(data, cutoff, fs, order=5):
    b, a = butter_lowpass(cutoff, fs, order=order)
    y = lfilter(b, a, data)
    return y

def seriesBWFilter(series, cutoff, order, fs = 1 / (60*default_sampling_interval)):
    return pd.Series( butter_lowpass_filter(series, cutoff=cutoff, order=order, fs=fs), index=series.index)

def colcom(c1, c2): #combine two equal-length lists
    cnew = [(i1+i2) /2 for i1,i2 in zip(c1,c2)]
    #print(cnew)
    return cnew
#some helpful wrappers...

        
def rename_units(PARC_ts, new_unit_name): #rename the units of a PARC_TimeSeries
                                            #note: doesn't fix tag label...
    PARC_ts.tag_units = new_unit_name
    PARC_ts.tag_description += " ("  + new_unit_name + ")"
def scale(PARC_ts, factor):#rescale the data of a PARC_TimeSeries
    PARC_ts.series *= factor
def title(ax, title):
    ax.set_title(title)
def comparison_plot(ax, PARC_ts1, PARC_ts2): #COULD be improved... move all series to average at zero,
                                             #then scale them to have stdDev of 1
    scalar = PARC_ts1.get_average() / PARC_ts2.get_average()
    PARC_ts1.plot(ax, scale=1, auto_title=False)
    PARC_ts2.plot(ax, scale=scalar, auto_title=False)
    title(ax,
            "COMPARISON RESCALE: " + PARC_ts1.tag + " v.s. " + PARC_ts2.tag + "\n" +
            PARC_ts1.tag_description  + " v.s. " + PARC_ts2.tag_description
          )
def comparison_plot_improved(PARC_ts_list, ax=None):
    newPARC_list = []
    for ts in PARC_ts_list:
        newPARC_list.append(ts.normalize())
        if ax != None:
            newPARC_list[-1].plot(ax)
    if ax!= None:
        title(ax,
                "COMPARISON RECENTER & RESCALE: " + PARC_ts_list[0].tag + " v.s. " + PARC_ts_list[1].tag + "\n" +
                PARC_ts_list[0].tag_description  + " v.s. " + PARC_ts_list[1].tag_description
              )
    return newPARC_list


def showplot():
    plt.show()

def qplot(PARCts, show=True):
    if type(PARCts) == list:
        if len(PARCts) > 1:
            n = len(PARCts)
            figq, ax = plt.subplots(n,1,sharex=True)
            for i,ts in enumerate(PARCts):
                if type(ts) == list:
                    for s in ts:
                        s.plot(ax[i])
                else:
                    ts.plot(ax[i])
            for a in ax:
                a.grid(True, 'both')
        if show: showplot()
        return
 
    figq, ax = plt.subplots()
    if type(PARCts) == list: #list of len 1
        for ts in PARCts: ts.plot(ax)
    else: PARCts.plot(ax)
    
    if show: showplot()
    return


def get_relationship(ts1, ts2):
    avg = [ts1.get_average(), ts2.get_average()]
    std = [ts1.get_standardDeviation(), ts2.get_standardDeviation()]
    # ts1 ~ (ts2 + off) * scale
    offset = avg[1] * std[0] / std[1] - avg[0]
    scalar = std[1] / std[0]
    print("\t",ts1.tag," VS ", ts2.tag)
    print("offset:",offset,"\nscalar:",scalar)
    return offset, scalar
 


mpl.rcParams['figure.figsize'] = (10,7)
mpl.rcParams['axes.titlepad'] = 0 # doesn't seem to do anything...
mpl.rcParams['axes.titlesize'] = 9
mpl.rcParams['figure.subplot.hspace'] = 0.5
mpl.rcParams['figure.subplot.top'] = 0.95
mpl.rcParams['figure.subplot.bottom'] = 0.05
'''
# READ EXCEL FILE

xlfile = "HD_K1_Tags.xlsx"

data = getData(xlfile)



#formatting
#define the figure
number_plots = 3
fig, ax = plt.subplots(number_plots,1,sharex=True)


datalist = [data[key] for key in data] #not necessary, but you can reference data by datalist[n]

speed = data["KAL.KAL_DeltaV.DRV25219 speed"]
blend = data["KAL.KAL_DeltaV.FFIC_K1_FL_BLEND OCC Flow SP"]
flow = data["KAL.KAL_DeltaV.FI212029-TPH"]

sxb = speed * blend

#do all of your calculations
comparison_plot_improved( ax[0], [speed, blend])
title(ax[0], "K1: Reel Speed and OCC Blend")
flow.plot(ax[1])
title(ax[1], "HD -> K1 flow")
comparison_plot_improved( ax[2], [flow,blend*flow])
title(ax[2], "HD -> K1 flow V.S. Speed x Blend")


showplot()

'''


'''
data = [PARC_TimeSeries(xlfile, "Raw Data "+str(i)) for i in range(1,5)]
for i, series in enumerate(data):
    series.plot(ax[i])

plt.show()

'''


'''

data1 = getTimeSeries(xlfile, "Raw Data 1") #get the data
print(data1)

data1diff = 10*seriesDerivative(data1) #take the derivative of the data


d1d_filt = seriesSVFilter(data1diff, 5, 2) #smooth out the slope

#graphing...
fig, ax = plt.subplots(3,1,sharex=True)

ax[0].plot(data1)

ax[1].plot( data1diff )

ax[2].plot( d1d_filt )
plt.show()
'''
